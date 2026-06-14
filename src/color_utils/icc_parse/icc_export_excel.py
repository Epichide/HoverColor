"""ICC 解析结果的 Excel 导出工具。

本模块负责把 ICC 解析后的 JSON 兼容结构导出为多 sheet 的 xlsx 文件。
GUI 只需要调用 export_to_excel，具体的数据整理、字段扁平化和表格样式都在这里完成。
"""

import json
from typing import Any, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PARSED_FIELD_REQUIRED_KEYS = {"value", "offset", "bytesize", "datatype"}
PARSED_FIELD_ALLOWED_KEYS = PARSED_FIELD_REQUIRED_KEYS | {"datasize"}
CONTAINER_METADATA_KEYS = {"offset", "bytesize", "datatype", "datasize"}
DETAIL_HEADERS = ["Tag", "Path", "Value", "Offset", "Bytesize", "Datatype", "Datasize"]


def export_to_excel(parsed_data: dict, output_path: str) -> None:
    """将 ICC 解析结果导出为多 sheet Excel 文件。

    Args:
        parsed_data: parse_icc_binary 或 JSON 导入得到的解析结果。
        output_path: 目标 .xlsx 文件路径。

    Returns:
        None. 文件会保存到 output_path。

    Example:
        export_to_excel(parsed_data, "profile_parsed.xlsx")
    """
    workbook = Workbook()

    basic_sheet = workbook.active
    basic_sheet.title = "Basic Info"
    _write_sheet(basic_sheet, ["Item", "Value"], _build_basic_rows(parsed_data))

    header_sheet = workbook.create_sheet("Header")
    _write_sheet(header_sheet, ["Item", "Value", "Offset", "Bytesize", "Datatype", "Datasize"], _build_header_rows(parsed_data))

    tags_sheet = workbook.create_sheet("Tags")
    _write_sheet(tags_sheet, ["Tag", "Type", "Offset", "Size", "Data Offset", "Data Size"], _build_tag_rows(parsed_data))

    details_sheet = workbook.create_sheet("Tag Details")
    _write_sheet(details_sheet, DETAIL_HEADERS, _build_tag_detail_rows(parsed_data))

    workbook.save(output_path)


def _build_basic_rows(parsed_data: dict) -> List[List[Any]]:
    """构造基础信息 sheet 的行数据。"""
    return [
        ["Parser", parsed_data.get("parser", "N/A")],
        ["File Path", parsed_data.get("file_path", "N/A")],
        ["File Size", f"{parsed_data.get('file_size', 0)} bytes"],
        ["Tag Count", parsed_data.get("tag_count", 0)],
    ]


def _build_header_rows(parsed_data: dict) -> List[List[Any]]:
    """构造 Header sheet 的行数据。"""
    rows = []
    header = parsed_data.get("header", {})
    for field_name, field_data in header.items():
        if isinstance(field_data, list) and len(field_data) >= 5:
            value, offset, bytesize, datatype, datasize = field_data[:5]
            rows.append([field_name, _format_cell_value(value), offset, bytesize, datatype, datasize])
        else:
            rows.append([field_name, _format_cell_value(field_data), "", "", "", ""])
    return rows


def _build_tag_rows(parsed_data: dict) -> List[List[Any]]:
    """构造 tag table sheet 的行数据。"""
    rows = []
    tags = parsed_data.get("tags", {})
    for tag_name in sorted(tags.keys()):
        tag_meta = tags.get(tag_name, {})
        if not isinstance(tag_meta, dict):
            rows.append([tag_name, "", "", "", "", ""])
            continue

        data_offset = _field_value(tag_meta.get("data_offset"))
        data_size = _field_value(tag_meta.get("data_size"))
        rows.append([
            tag_name,
            tag_meta.get("type", ""),
            tag_meta.get("offset", ""),
            tag_meta.get("size", ""),
            data_offset,
            data_size,
        ])
    return rows


def _build_tag_detail_rows(parsed_data: dict) -> List[List[Any]]:
    """构造所有 tag 解析详情的扁平行数据。"""
    rows = []
    tag_data = parsed_data.get("tag_data", {})
    for tag_name in sorted(tag_data.keys()):
        rows.extend(_flatten_tag_fields(tag_name, tag_data.get(tag_name), ""))
    return rows


def _flatten_tag_fields(tag_name: str, data: Any, path: str) -> List[List[Any]]:
    """递归展开一个 tag 的嵌套字段，保留字段路径和解析元数据。"""
    rows = []

    if _is_parsed_field(data):
        rows.append([
            tag_name,
            path or "value",
            _format_cell_value(data.get("value")),
            data.get("offset", ""),
            data.get("bytesize", ""),
            data.get("datatype", ""),
            data.get("datasize", ""),
        ])
        return rows

    if isinstance(data, dict):
        if path:
            rows.append([
                tag_name,
                path,
                "",
                data.get("offset", ""),
                data.get("bytesize", ""),
                data.get("datatype", ""),
                data.get("datasize", ""),
            ])

        for key, value in data.items():
            if key in CONTAINER_METADATA_KEYS:
                continue
            child_path = f"{path}.{key}" if path else key
            rows.extend(_flatten_tag_fields(tag_name, value, child_path))
        return rows

    if isinstance(data, list):
        if _is_simple_list(data):
            rows.append([tag_name, path, _format_cell_value(data), "", "", "", len(data)])
        else:
            rows.append([tag_name, path, f"[{len(data)} items]", "", "", "", len(data)])
            for index, item in enumerate(data):
                rows.extend(_flatten_tag_fields(tag_name, item, f"{path}[{index}]"))
        return rows

    rows.append([tag_name, path or "value", _format_cell_value(data), "", "", "", ""])
    return rows


def _write_sheet(worksheet, headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> None:
    """写入一个 sheet，并应用基础样式、筛选和列宽。"""
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append([_format_cell_value(value) for value in row])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 80)


def _is_parsed_field(value: Any) -> bool:
    """判断是否为单个解析字段包装。"""
    if not isinstance(value, dict):
        return False
    return PARSED_FIELD_REQUIRED_KEYS.issubset(value.keys()) and set(value.keys()).issubset(PARSED_FIELD_ALLOWED_KEYS)


def _field_value(value: Any, default: Any = "") -> Any:
    """兼容裸值和 ParsedField 结构，取出真实值。"""
    if _is_parsed_field(value):
        return value.get("value", default)
    return value if value is not None else default


def _is_simple_list(value: list) -> bool:
    """判断列表是否可以直接摘要到一个单元格中。"""
    return all(isinstance(item, (str, int, float, bool)) or item is None for item in value)


def _format_cell_value(value: Any, max_items: int = 24, max_chars: int = 32000) -> str:
    """格式化 Excel 单元格值，避免超出 Excel 单元格长度限制。"""
    if value is None:
        return ""

    if isinstance(value, list):
        if _is_simple_list(value):
            preview = ", ".join(str(item) for item in value[:max_items])
            if len(value) > max_items:
                preview = f"{preview}, ... ({len(value)} items)"
            text = f"[{preview}]"
        else:
            text = json.dumps(value[:max_items], ensure_ascii=False)
            if len(value) > max_items:
                text = f"{text[:-1]}, ... ({len(value)} items)]"
    elif isinstance(value, dict):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)

    if len(text) > max_chars:
        return text[:max_chars - 32] + f"... ({len(text)} chars)"
    return text
